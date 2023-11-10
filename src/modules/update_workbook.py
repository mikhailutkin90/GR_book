import csv
import openpyxl
import dropbox
from openpyxl.styles import Border, Side, PatternFill



def update_workbook(file_name_entry):
    '''Updates the bookkeeping workbook with data entered by user in GUI'''
    cum_input_file = "/" + file_name_entry.get() + "_cum_input.csv"
    with open('key.txt', 'r') as file:
        key = file.read()

    dbx = dropbox.Dropbox(key)
    _, response = dbx.files_download(cum_input_file)
    csv_content = response.content.decode('utf-8')
    
    csv_reader = csv.reader(csv_content.splitlines())
    next(csv_reader)
    row_count=0
    for row_c in csv_reader:
        row_count += 1    
    
    transaction_dates = [0] * (row_count)
    transaction_names = [0] * (row_count)
    transaction_ammounts = [0] * (row_count)
    for_manual_check = [0] * (row_count)
    vat = [0] * (row_count)
    codes = [0] * (row_count)

    csv_reader = csv.reader(csv_content.splitlines())
    next(csv_reader)
    i = 0
    for row in csv_reader:
        transaction_dates[i] = row[1]
        transaction_names[i] = row[2]
        transaction_ammounts[i] = float(row[3])
        for_manual_check[i] = int(row[6])
        codes[i] = int(row[7])
        if int(row[4]) == 1:
            vat[i] = 24
        if int(row[5]) == 1:
            vat[i] = 0
        i += 1 
   

    workbook = openpyxl.load_workbook('bookkeep_table.xlsx')
    worksheet = workbook.active
    for it in range(row_count):
        if for_manual_check[it] == 1:
            last_row = worksheet.max_row
            #calculation of the new transaction PT - value
            PT_old_value = worksheet.cell(row=(last_row-4), column=2).value
            PT_new_value = PT_old_value[:-3] + str(int(PT_old_value[-3:]) + 1)
           
            row1 = ('TO BE CHECKED MANUALLY ', ' ', ' ', ' ', ' ', ' ', ' ')
            row2 = ('date', 'voucher', ' ', 'vat%', 'description', 'debit', 'credit')
            row3 = (transaction_dates[it], PT_new_value, '  ', '  ', transaction_names[it], '  ', '  ')
            row4 = (' ', 0000, ' ', 0, ' ', 0, abs(round(transaction_ammounts[it], 2)))
            row5 = (' ', 0000, '___', 0, ' ', 0 , 0)
            row6 = (' ', 0000, ' ', 0, ' ', 0, 0)
            row7 = (' ', ' ', ' ', ' ', 'Total', 0, 0)

            worksheet.append(row1)
            worksheet.append(row2)
            worksheet.append(row3)
            worksheet.append(row4)
            worksheet.append(row5)
            worksheet.append(row6)
            worksheet.append(row7)
            last_row = worksheet.max_row
            thick_border = Border(bottom=Side(style='thick'))
            for cell in worksheet[(last_row-1)]:
                cell.border = thick_border
            for cell in worksheet[(last_row-5)]:
                cell.border = thick_border
            last_row_index = worksheet.max_row
            # Apply red background color to the last 7 rows
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            for row in worksheet.iter_rows(min_row=last_row_index - 6, max_row=last_row_index, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.fill = red_fill

        if transaction_ammounts[it] < 0 and for_manual_check[it] == 0:
            last_row = worksheet.max_row
            #calculation of the new transaction PT - value
            PT_old_value = worksheet.cell(row=(last_row-4), column=2).value
            PT_new_value = PT_old_value[:-3] + str(int(PT_old_value[-3:]) + 1)
           
            row1 = (' ', ' ', ' ', ' ', ' ', ' ', ' ')
            row2 = ('date', 'voucher', ' ', 'vat%', 'description', 'debit', 'credit')
            row3 = (transaction_dates[it], PT_new_value, '  ', '  ', transaction_names[it], '  ', '  ')
            #1910 is a standard code for all credit operations from companies account
            row4 = (' ', 1910, ' ', 0, ' ', 0, (-1) * transaction_ammounts[it])
            without_vat_value = round(row4[6] / (1+vat[it]/100), 2)
            row5 = (' ', codes[it], '___', vat[it], ' ', without_vat_value , 0)
            vat_value = round((row4[6] - row5[5]), 2)
            #1763 is a standard code for all vat recieved to the company
            row6 = (' ', 1763, 'VAT received', 0, ' ', vat_value, 0)
            row7 = (' ', ' ', ' ', ' ', 'Total', (row4[5]+row5[5]+row6[5]), (row4[6]+row5[6]+row6[6]))
            worksheet.append(row1)
            worksheet.append(row2)
            worksheet.append(row3)
            worksheet.append(row4)
            worksheet.append(row5)
            worksheet.append(row6)
            worksheet.append(row7)

            last_row = worksheet.max_row

            thick_border = Border(bottom=Side(style='thick'))
            for cell in worksheet[(last_row-1)]:
                cell.border = thick_border
            for cell in worksheet[(last_row-5)]:
                cell.border = thick_border

        if transaction_ammounts[it] > 0 and for_manual_check[it] == 0:
            last_row = worksheet.max_row
            #calculation of the new transaction PT - value
            PT_old_value = worksheet.cell(row=(last_row-4), column=2).value
            PT_new_value = PT_old_value[:-3] + str(int(PT_old_value[-3:]) + 1)
            row1 = (' ', ' ', ' ', ' ', ' ', ' ', ' ')
            row2 = ('date', 'voucher', ' ', 'vat%', 'description', 'debit', 'credit')
            row3 = (transaction_dates[it], PT_new_value, '  ', '  ', transaction_names[it], '  ', '  ')
            row4 = (' ', 3000, 'Sales ', 0, ' ', 0, round(transaction_ammounts[it]/(1+vat[it]/100), 2))
            row5 = (' ', 1701, 'Trade debtors', vat[it], ' ', round(transaction_ammounts[it], 2) , 0)
            vat_value = round((row5[5] - row4[6]), 2)
            row6 = (' ', 2939, 'VAT liability ', 0, ' ', 0, vat_value)
            row7 = (' ', ' ', ' ', ' ', 'Total', (row4[5]+row5[5]+row6[5]), (row4[6]+row5[6]+row6[6]))

            # Append the rows to the worksheet
            worksheet.append(row1)
            worksheet.append(row2)
            worksheet.append(row3)
            worksheet.append(row4)
            worksheet.append(row5)
            worksheet.append(row6)
            worksheet.append(row7)

            last_row = worksheet.max_row

            thick_border = Border(bottom=Side(style='thick'))
            for cell in worksheet[(last_row-1)]:
                cell.border = thick_border
            for cell in worksheet[(last_row-5)]:
                cell.border = thick_border
           
    workbook.save('bookkeep_table.xlsx')
   

